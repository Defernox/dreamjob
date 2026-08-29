"""Export et reprise du suivi de candidatures au format Excel.

Le fichier produit est destiné à être **envoyé tel quel à France Travail** comme
justificatif de recherche d'emploi : une ligne par candidature, des en-têtes
lisibles en français, des dates au format jj/mm/aaaa, la ligne de titre figée.
Aucune formule, aucun onglet technique — un agent doit pouvoir le lire sans
explication.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models.enums import STATUTS

log = logging.getLogger("dreamjob.excel")

FEUILLE = "Candidatures"

# (en-tête, clé, largeur). L'ordre est celui du cahier des charges.
COLONNES: list[tuple[str, str, int]] = [
    ("Date de candidature", "date_candidature", 20),
    ("Entreprise", "entreprise", 28),
    ("Poste", "titre", 42),
    ("Pays", "pays", 14),
    ("Score", "score", 8),
    ("Date limite", "deadline", 14),
    ("Statut", "statut", 14),
    ("Notes", "notes", 40),
    ("Contact", "contact", 24),
    ("Lien de l'offre", "url", 46),
]

_ENTETE_FOND = PatternFill("solid", fgColor="1F2937")
_ENTETE_POLICE = Font(color="FFFFFF", bold=True, size=11)
_BORDURE = Border(bottom=Side(style="thin", color="D1D5DB"))


def _en_date(valeur) -> date | None:
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    return None


def exporter(candidatures: list[dict]) -> bytes:
    """Classeur .xlsx en mémoire, prêt à être téléchargé."""
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = FEUILLE

    for numero, (entete, _, largeur) in enumerate(COLONNES, start=1):
        cellule = feuille.cell(row=1, column=numero, value=entete)
        cellule.fill = _ENTETE_FOND
        cellule.font = _ENTETE_POLICE
        cellule.alignment = Alignment(vertical="center")
        feuille.column_dimensions[get_column_letter(numero)].width = largeur
    feuille.row_dimensions[1].height = 22

    for ligne, candidature in enumerate(candidatures, start=2):
        for numero, (_, cle, _) in enumerate(COLONNES, start=1):
            valeur = candidature.get(cle)
            cellule = feuille.cell(row=ligne, column=numero)

            if cle in ("date_candidature", "deadline"):
                jour = _en_date(valeur)
                cellule.value = jour
                if jour is not None:
                    cellule.number_format = "DD/MM/YYYY"
            elif cle == "score":
                cellule.value = round(valeur) if valeur is not None else None
            else:
                cellule.value = valeur or ""

            cellule.alignment = Alignment(
                vertical="top", wrap_text=cle in ("notes", "titre")
            )
            cellule.border = _BORDURE

    # En-têtes figés + filtres : le fichier reste lisible avec 200 lignes.
    feuille.freeze_panes = "A2"
    feuille.auto_filter.ref = (
        f"A1:{get_column_letter(len(COLONNES))}{max(len(candidatures) + 1, 1)}"
    )

    tampon = BytesIO()
    classeur.save(tampon)
    log.info("Export Excel : %d candidatures", len(candidatures))
    return tampon.getvalue()


# ------------------------------------------------------------------- reprise


@dataclass
class Reprise:
    """Compte rendu d'un import : ce qui a été repris, ce qui n'a pas pu l'être."""

    mises_a_jour: int = 0
    ignorees: int = 0
    problemes: list[str] = field(default_factory=list)
    lignes: list[dict] = field(default_factory=list)


def lire(contenu: bytes | Path) -> Reprise:
    """Relit un export et en extrait les colonnes de suivi.

    Tolérant sur l'ordre des colonnes : on se repère aux en-têtes, pas aux
    positions — un fichier retouché à la main reste importable.
    """
    source = BytesIO(contenu) if isinstance(contenu, bytes) else contenu
    try:
        classeur = load_workbook(source, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"Fichier Excel illisible : {e}") from e

    feuille = classeur[FEUILLE] if FEUILLE in classeur.sheetnames else classeur.active
    rangees = list(feuille.iter_rows(values_only=True))
    if not rangees:
        raise ValueError("Le fichier est vide.")

    entetes = [str(c).strip() if c is not None else "" for c in rangees[0]]
    par_entete = {entete: cle for entete, cle, _ in COLONNES}
    indices = {par_entete[e]: i for i, e in enumerate(entetes) if e in par_entete}

    manquantes = {"entreprise", "titre"} - set(indices)
    if manquantes:
        raise ValueError(
            "Colonnes indispensables absentes : "
            + ", ".join(e for e, c, _ in COLONNES if c in manquantes)
        )

    reprise = Reprise()
    for numero, rangee in enumerate(rangees[1:], start=2):
        def valeur(cle):
            i = indices.get(cle)
            return rangee[i] if i is not None and i < len(rangee) else None

        entreprise = (valeur("entreprise") or "").strip()
        titre = (valeur("titre") or "").strip()
        if not entreprise and not titre:
            continue      # ligne vide en fin de feuille

        statut = (valeur("statut") or "").strip()
        if statut and statut not in STATUTS:
            reprise.problemes.append(f"Ligne {numero} : statut inconnu « {statut} », ignoré.")
            statut = ""

        reprise.lignes.append({
            "entreprise": entreprise,
            "titre": titre,
            "url": (valeur("url") or "").strip(),
            "statut": statut,
            "notes": (valeur("notes") or "").strip(),
            "contact": (valeur("contact") or "").strip(),
            "deadline": _en_date(valeur("deadline")),
            "ligne": numero,
        })
    return reprise
