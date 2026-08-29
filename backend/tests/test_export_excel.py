"""L'export Excel — le justificatif envoyé à France Travail.

Il doit se lire sans explication : en-têtes en français, dates jj/mm/aaaa, une
ligne par candidature, ligne de titre figée.
"""

from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlmodel import Session

from app.exports.excel import COLONNES, exporter, lire
from app.models import Application, Offer


def _donnee(**kw) -> dict:
    base = dict(
        date_candidature=datetime(2026, 8, 29, 14, 30), entreprise="Assureur Crédit",
        titre="Chargé de recouvrement (H/F)", pays="France", score=82.4,
        deadline=date(2026, 9, 15), statut="Envoyée", notes="Relancer lundi.",
        contact="rh@exemple.fr", url="https://exemple.test/offre/1",
    )
    return {**base, **kw}


def _feuille(octets: bytes):
    return load_workbook(BytesIO(octets)).active


# --- Forme du fichier -------------------------------------------------------


def test_les_entetes_sont_en_francais_et_dans_l_ordre_attendu():
    feuille = _feuille(exporter([_donnee()]))
    entetes = [c.value for c in feuille[1]]
    assert entetes == [e for e, _, _ in COLONNES]
    assert entetes[0] == "Date de candidature"


def test_une_ligne_par_candidature():
    feuille = _feuille(exporter([_donnee(), _donnee(entreprise="Banque B")]))
    assert feuille.max_row == 3      # 1 en-tête + 2 candidatures


def test_les_dates_sont_au_format_francais():
    feuille = _feuille(exporter([_donnee()]))
    cellule = feuille.cell(row=2, column=1)
    assert cellule.number_format == "DD/MM/YYYY"
    # Une vraie date, pas du texte : Excel doit pouvoir trier dessus.
    assert isinstance(cellule.value, (date, datetime))


def test_la_ligne_de_titre_est_figee():
    """Sans cela, un agent qui fait défiler 200 lignes perd les en-têtes."""
    assert _feuille(exporter([_donnee()])).freeze_panes == "A2"


def test_le_score_est_arrondi():
    feuille = _feuille(exporter([_donnee(score=82.4)]))
    assert feuille.cell(row=2, column=5).value == 82


def test_un_export_vide_reste_un_fichier_valide():
    feuille = _feuille(exporter([]))
    assert feuille.max_row == 1
    assert feuille.cell(row=1, column=1).value == "Date de candidature"


def test_les_champs_absents_donnent_une_cellule_vide_pas_le_mot_none():
    """Un « None » écrit en toutes lettres dans un justificatif officiel ferait
    mauvais effet. Une cellule vide (None côté openpyxl) s'affiche blanche : c'est
    le texte « None » qu'il faut proscrire."""
    feuille = _feuille(exporter([_donnee(notes=None, contact=None, deadline=None)]))
    textes = [c.value for c in feuille[2] if isinstance(c.value, str)]
    assert "None" not in textes
    assert feuille.cell(row=2, column=6).value is None      # date limite vide


# --- Reprise ----------------------------------------------------------------


def test_aller_retour_complet():
    reprise = lire(exporter([_donnee()]))
    ligne = reprise.lignes[0]
    assert ligne["entreprise"] == "Assureur Crédit"
    assert ligne["statut"] == "Envoyée"
    assert ligne["deadline"] == date(2026, 9, 15)
    assert ligne["notes"] == "Relancer lundi."


def test_l_ordre_des_colonnes_peut_changer():
    """Un fichier retouché à la main doit rester importable."""
    from openpyxl import Workbook

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Candidatures"
    feuille.append(["Poste", "Statut", "Entreprise"])
    feuille.append(["Analyste risques", "Entretien", "Banque A"])
    tampon = BytesIO()
    classeur.save(tampon)

    ligne = lire(tampon.getvalue()).lignes[0]
    assert ligne["entreprise"] == "Banque A"
    assert ligne["statut"] == "Entretien"


def test_un_statut_inconnu_est_signale_et_ignore():
    from openpyxl import Workbook

    classeur = Workbook()
    feuille = classeur.active
    feuille.append(["Entreprise", "Poste", "Statut"])
    feuille.append(["Banque A", "Analyste", "Peut-être"])
    tampon = BytesIO()
    classeur.save(tampon)

    reprise = lire(tampon.getvalue())
    assert reprise.lignes[0]["statut"] == ""
    assert any("Peut-être" in p for p in reprise.problemes)


def test_les_colonnes_indispensables_sont_exigees():
    from openpyxl import Workbook

    classeur = Workbook()
    classeur.active.append(["Statut", "Notes"])
    tampon = BytesIO()
    classeur.save(tampon)

    with pytest.raises(ValueError, match="Entreprise"):
        lire(tampon.getvalue())


def test_un_fichier_illisible_le_dit():
    with pytest.raises(ValueError, match="illisible"):
        lire(b"ceci n'est pas un classeur")


# --- Endpoints --------------------------------------------------------------


@pytest.fixture
def candidature(client, engine):
    with Session(engine) as s:
        offre = Offer(source="france_travail", source_id="1", hash="h1",
                      titre="Analyste risques de crédit", entreprise="Banque A",
                      pays="France", score=88.0, url="https://exemple.test/offre/1")
        s.add(offre)
        s.commit()
        s.refresh(offre)
        identifiant = offre.id          # lu AVANT la fermeture de la session
        s.add(Application(offer_id=identifiant, notes="À relancer"))
        s.commit()
    return identifiant


def test_telechargement_du_fichier(client, candidature):
    reponse = client.get("/api/candidatures/export.xlsx")
    assert reponse.status_code == 200
    assert "spreadsheetml" in reponse.headers["content-type"]
    assert "candidatures-" in reponse.headers["content-disposition"]

    feuille = _feuille(reponse.content)
    assert feuille.cell(row=2, column=2).value == "Banque A"
    assert feuille.cell(row=2, column=3).value == "Analyste risques de crédit"


def test_reprise_met_a_jour_le_suivi(client, candidature):
    """Le cas d'usage : on retouche le fichier dans Excel, on le réimporte."""
    export = client.get("/api/candidatures/export.xlsx").content
    classeur = load_workbook(BytesIO(export))
    feuille = classeur.active
    entetes = [c.value for c in feuille[1]]
    feuille.cell(row=2, column=entetes.index("Statut") + 1, value="Entretien")
    feuille.cell(row=2, column=entetes.index("Notes") + 1, value="Entretien le 12/09")
    tampon = BytesIO()
    classeur.save(tampon)

    resultat = client.post(
        "/api/candidatures/importer",
        files={"fichier": ("suivi.xlsx", tampon.getvalue(),
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()
    assert resultat["mises_a_jour"] == 1

    candidatures = client.get("/api/candidatures").json()
    assert candidatures[0]["statut"] == "Entretien"
    assert candidatures[0]["notes"] == "Entretien le 12/09"


def test_une_ligne_sans_offre_correspondante_est_signalee(client, candidature):
    from openpyxl import Workbook

    classeur = Workbook()
    classeur.active.append(["Entreprise", "Poste", "Statut"])
    classeur.active.append(["Société Inconnue", "Poste fantôme", "Refus"])
    tampon = BytesIO()
    classeur.save(tampon)

    resultat = client.post(
        "/api/candidatures/importer",
        files={"fichier": ("suivi.xlsx", tampon.getvalue(), "application/vnd.ms-excel")},
    ).json()
    assert resultat["mises_a_jour"] == 0
    assert resultat["ignorees"] == 1
    assert "Société Inconnue" in resultat["problemes"][0]


def test_un_format_non_excel_est_refuse(client):
    reponse = client.post(
        "/api/candidatures/importer",
        files={"fichier": ("suivi.csv", b"a,b,c", "text/csv")},
    )
    assert reponse.status_code == 400
